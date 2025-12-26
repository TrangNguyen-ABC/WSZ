import openpyxl
from playwright.sync_api import sync_playwright
import time

# --- CẤU HÌNH ---
INPUT_FILE = 'check_task_hero/test_1.xlsx'  # Tên file Excel đầu vào
BASE_URL = "https://worksheetzone.org/" # Đường dẫn gốc
SHEET_NAME = None # Để None nếu muốn lấy sheet đầu tiên, hoặc điền tên sheet 'Sheet1'

def process_excel():
    # 1. Mở file Excel
    try:
        wb = openpyxl.load_workbook(INPUT_FILE)
        if SHEET_NAME:
            sheet = wb[SHEET_NAME]
        else:
            sheet = wb.active # Lấy sheet đang active
    except FileNotFoundError:
        print(f"Lỗi: Không tìm thấy file '{INPUT_FILE}'. Hãy kiểm tra lại tên file.")
        return

    # 2. Khởi động Playwright
    with sync_playwright() as p:
        # Mở trình duyệt (headless=True để chạy ẩn cho nhanh)
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
        page = context.new_page()

        print(f"Bắt đầu xử lý file: {INPUT_FILE}...")
        
        # 3. Duyệt qua các hàng trong cột A
        # iter_rows trả về tuple các cell, ở đây ta lấy max_col=1 nghĩa là chỉ lấy cột A
        row_index = 0
        for row in sheet.iter_rows(min_row=1, max_col=1):
            cell_slug = row[0] # Ô dữ liệu cột A
            slug = cell_slug.value
            
            # Bỏ qua nếu ô trống
            if not slug:
                continue
            
            row_index += 1
            # Ghép URL (đảm bảo xử lý dấu / nếu cần thiết)
            # Logic: https://worksheetzone.org/ + slug
            # Lưu ý: Nếu slug trong file excel có chứa dấu / ở đầu hoặc base_url thiếu dấu /, hãy điều chỉnh ở đây.
            full_url = f"{BASE_URL.rstrip('/')}/{str(slug).strip('/')}"
            
            print(f"[{row_index}] Checking: {full_url}", end=" -> ")

            status = "Error" # Mặc định là lỗi nếu try/except bắt được
            
            try:
                # Truy cập trang web
                # wait_until='domcontentloaded' giúp load nhanh hơn thay vì đợi load toàn bộ ảnh
                page.goto(full_url, wait_until='domcontentloaded', timeout=10000) 
                
                # Tìm thẻ meta robots
                meta_robots = page.locator('meta[name="robots"]')
                
                # Logic kiểm tra
                if meta_robots.count() > 0:
                    content = meta_robots.first.get_attribute("content")
                    if content and "noindex" in content.lower():
                        status = "FAIL" # Có noindex -> Fail
                    else:
                        status = "--" # Có thẻ robots nhưng không có noindex -> Pass
                else:
                    status = "--" # Không có thẻ robots -> Mặc định Google sẽ index -> Pass
                
            except Exception as e:
                print(f"(Lỗi: {str(e)})", end=" ")
                status = "Error" # Lỗi mạng hoặc timeout

            print(status)

            # 4. Ghi kết quả vào cột B (cột 2) cùng dòng đó
            # row[0].row là số thứ tự dòng hiện tại
            sheet.cell(row=cell_slug.row, column=2).value = status

        # Đóng trình duyệt
        browser.close()

    # 5. Lưu file Excel lại
    try:
        wb.save(INPUT_FILE)
        print(f"\nĐã hoàn thành! Kết quả đã được lưu vào cột B trong file '{INPUT_FILE}'.")
    except PermissionError:
        print(f"\nLỖI: Không thể lưu file. Hãy tắt file Excel '{INPUT_FILE}' trước khi chạy code.")

if __name__ == "__main__":
    process_excel()